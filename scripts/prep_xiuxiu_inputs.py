#!/usr/bin/env python3
"""咻咻勇者输入预处理（不使用术语表，两份都是待提取源）：
1. 美术字统计：合并 4 sheet 的「简体」列 -> 单列源 source_art.xlsx
2. LanguageDesc：跳过元数据行（row 0-8 = 表头/类型声明/语言标签），取 ZH+EN -> 干净双语源 source_lang_bilingual.xlsx
3. 生成空术语表 empty_glossary.xlsx（流水线必须传 --glossary，空表=不参与匹配）
"""
import argparse
from pathlib import Path
import openpyxl
import pandas as pd

_DATA = Path("/Users/spellbook/Desktop/Langlobal/杂项文档")
ap = argparse.ArgumentParser(description="咻咻勇者输入预处理")
ap.add_argument("--art", type=Path, default=_DATA / "台服美术字统计_F列简中译英.xlsx",
                help="美术字统计 xlsx")
ap.add_argument("--lang", type=Path, default=_DATA / "LanguageDesc.xlsx",
                help="LanguageDesc xlsx")
ap.add_argument("--out", type=Path, default=_DATA / "xiuxiu_prep",
                help="产物输出目录")
args = ap.parse_args()

OUT = args.out
OUT.mkdir(parents=True, exist_ok=True)

# ── 1. 美术字 -> 单列源 ──
src = args.art
wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
texts = []
for ws in wb.worksheets:
    rows = ws.iter_rows(values_only=True)
    hdr = [("" if c is None else str(c)) for c in next(rows)]
    if "简体" not in hdr:
        print(f"[warn] sheet {ws.title} 无『简体』列，跳过"); continue
    idx = hdr.index("简体")
    for r in rows:
        v = r[idx] if idx < len(r) else None
        if v is not None and str(v).strip():
            texts.append(str(v).strip())
wb.close()
uniq = list(dict.fromkeys(texts))
pd.DataFrame({"简体": uniq}).to_excel(OUT / "source_art.xlsx", index=False)
print(f"美术字源: {len(texts)} 行 -> 去重 {len(uniq)} -> {OUT/'source_art.xlsx'}")

# ── 2. LanguageDesc -> 干净双语源 (中文, 英文) ──
# row 0-8 是表头元数据（类型声明、语言标签等），直接跳过，不靠规则猜
SKIP_ROWS = frozenset(range(9))
lang = args.lang
wb = openpyxl.load_workbook(lang, read_only=True, data_only=True)
ws = wb["Sheet1"]
pairs = []
for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i in SKIP_ROWS:
        continue
    v = [("" if c is None else str(c)).strip() for c in r[:5]]
    key = v[1] if len(v) > 1 else ""
    zh = v[3] if len(v) > 3 else ""
    en = v[4] if len(v) > 4 else ""
    if key.startswith("##") or not key:
        continue
    if zh:
        pairs.append((zh, en))
wb.close()
dedup = {}
for zh, en in pairs:
    dedup.setdefault(zh, en)
pd.DataFrame([(k, v) for k, v in dedup.items()], columns=["中文", "英文"]).to_excel(
    OUT / "source_lang_bilingual.xlsx", index=False)
print(f"LanguageDesc 双语源: {len(pairs)} 行 -> 去重 {len(dedup)} -> {OUT/'source_lang_bilingual.xlsx'}")

# ── 3. 空术语表 ──
pd.DataFrame(columns=["中文", "英文"]).to_excel(OUT / "empty_glossary.xlsx", index=False)
print(f"空术语表 -> {OUT/'empty_glossary.xlsx'}")
