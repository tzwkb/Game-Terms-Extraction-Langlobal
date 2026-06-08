"""Generate the project-info intake form for PMs (optimized per Profile 撰写指南).
AI later parses the filled form into profiles/<game>.yaml.
4 columns: 问题 | 填写提示 | 参考示例(燕云) | 请在这里填写.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUT = "/Users/spellbook/Desktop/Langlobal/AI Lab/Game-Terms-Extraction-Langlobal/docs/项目配置填写表.xlsx"
LAST_COL = 4  # A..D

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
STAR_FILL    = PatternFill("solid", fgColor="FBF3DF")  # glossary / high-impact highlight
INPUT_FILL   = PatternFill("solid", fgColor="FFFDE7")  # yellow = type here
HINT_FILL    = PatternFill("solid", fgColor="F5F5F5")  # gray = hint
EXAMPLE_FILL = PatternFill("solid", fgColor="EAF3FB")  # light blue = reference example
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=12)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT   = Font(bold=True, size=10)
STAR_FONT    = Font(bold=True, size=10, color="B07D12")
HINT_FONT    = Font(italic=True, color="888888", size=9)
EX_FONT      = Font(color="1B5E9A", size=9)
BODY_FONT    = Font(size=10)
NOTE_FONT    = Font(color="7A560C", size=9)

BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP   = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply(cell, fill=None, font=None, align=WRAP, border=BORDER):
    if fill:   cell.fill = fill
    if font:   cell.font = font
    if align:  cell.alignment = align
    if border: cell.border = border


def banner(ws, row, text, fill, font, height, align=CENTER):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
    c = ws.cell(row=row, column=1, value=text)
    apply(c, fill=fill, font=font, align=align)
    for col in range(2, LAST_COL + 1):
        apply(ws.cell(row=row, column=col), fill=fill, border=BORDER)
    ws.row_dimensions[row].height = height


def section_header(ws, row, text):
    banner(ws, row, text, SECTION_FILL, SECTION_FONT, 24)


def checklist_row(ws, row, label, desc, height, star=False):
    """col1 = item label; col2..D merged = description."""
    lc = ws.cell(row=row, column=1, value=label)
    apply(lc, fill=(STAR_FILL if star else WHITE_FILL), font=(STAR_FONT if star else LABEL_FONT))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=LAST_COL)
    dc = ws.cell(row=row, column=2, value=desc)
    apply(dc, fill=(STAR_FILL if star else HINT_FILL), font=(NOTE_FONT if star else BODY_FONT))
    for col in range(3, LAST_COL + 1):
        apply(ws.cell(row=row, column=col), fill=(STAR_FILL if star else HINT_FILL), border=BORDER)
    ws.row_dimensions[row].height = height


def question_row(ws, row, label, hint, example, height=40, star=False):
    lc = ws.cell(row=row, column=1, value=label)
    apply(lc, fill=(STAR_FILL if star else WHITE_FILL), font=(STAR_FONT if star else LABEL_FONT))
    apply(ws.cell(row=row, column=2, value=hint), fill=HINT_FILL, font=HINT_FONT)
    apply(ws.cell(row=row, column=3, value=example), fill=EXAMPLE_FILL, font=EX_FONT)
    apply(ws.cell(row=row, column=4, value=""), fill=INPUT_FILL, font=BODY_FONT)
    ws.row_dimensions[row].height = height


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目信息填写表"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 46

    # ── Title + guide ─────────────────────────────────────────────
    banner(ws, 1, "游戏术语抽取 — 项目信息填写表", HEADER_FILL, HEADER_FONT, 36)
    banner(ws, 2,
           "🟡 只填最后一列（黄色）  |  第三列是燕云项目参考示例，照着写  |  ★ 题最影响准确率，请重点写  |  详见配套《Profile撰写指南.html》",
           HINT_FILL, HINT_FONT, 20)

    row = 3

    # ── 第零部分：开工前先备齐资料 ────────────────────────────────
    section_header(ws, row, "零、开工前先备齐这些资料（不用填，自查清单）")
    row += 1
    checklist_row(ws, row,
        "☐ 源文件",
        "游戏文本 .xlsx，至少一列是中文原文。变量占位符（{0}、${name}、<color> 等）自动剥离，不用手动清。可多文件分批（剧情/UI/技能拆开更好管）。",
        58)
    row += 1
    checklist_row(ws, row,
        "☐ 术语表 ★命门",
        "已有「中文术语 | 英文翻译」两列的 .xlsx。它是准确率命门：命中词被保护不误删、直接套用已有译文、并作新词近义参考——实测约九成召回靠它。"
        "来源：官方词表 / 历史项目 TM / 人物表 / 技能表 / 道具表，能并就并，越全越准。没有？先凑核心人物+核心系统词的小表垫着，跑完把确认的术语回灌，边跑边长。",
        96, star=True)
    row += 1
    checklist_row(ws, row,
        "☐ 填表素材",
        "世界观设定、人物/技能/道具/地名表、游戏系统结构、试译反馈、客户 brief——用来把下面 12 题填准。",
        46)
    row += 1

    # ── 栏头 ──────────────────────────────────────────────────────
    for col, text in [(1, "问题"), (2, "填写提示"), (3, "✍ 参考示例（燕云）"), (4, "⬇ 请在这里填写")]:
        apply(ws.cell(row=row, column=col, value=text), fill=SECTION_FILL, font=SECTION_FONT, align=CENTER)
    ws.row_dimensions[row].height = 22
    row += 1

    # ── 一、基本信息 ──────────────────────────────────────────────
    section_header(ws, row, "一、基本信息"); row += 1
    question_row(ws, row, "1. 游戏名称", "游戏的完整中文名称。",
        "燕云十六州", 30); row += 1
    question_row(ws, row, "2. 游戏风格 / 世界观",
        "一两句定调，模型据此定身份。\n例如：古风武侠、西幻奇幻、科幻末日、现代都市……",
        "古风武侠开放世界，五代乱世背景，含墨家机关术。", 48); row += 1
    question_row(ws, row, "3. 本次任务说明",
        "一句话说明用途。",
        "抽取游戏内所有专有名词，供翻译团队使用。", 40); row += 1

    # ── 二、要提取什么 ────────────────────────────────────────────
    section_header(ws, row, "二、要提取的内容（告诉 AI 抓什么 · 影响最大）"); row += 1
    question_row(ws, row, "4. ★ 必须提取的术语类型",
        "逐类列全，越详细越好。\n"
        "⚠ 关键：用「出现即提取，不论主次」这种绝对口吻，别用「重要的/酌情」——"
        "系统要 3 轮一致才保留，口径越绝对越不漏。",
        "· 所有人名/NPC名（含路人、单次提及，出现即提取）\n"
        "· 武学招式 / 心法\n· 墨家机关术\n· 地名建筑、门派势力\n"
        "· 道具圣物、货币资源\n· 战斗 BUFF / 机制、UI 系统、玩法机制",
        150, star=True); row += 1
    question_row(ws, row, "5. ★ 不要提取的内容",
        "列出噪音类型，这是降噪主力。",
        "· 通用日常物品（木炭、草鞋、柴火）\n"
        "· 泛称/称谓（大侠、公子、弟子、长老）\n"
        "· 无专名场所（书房、医馆、夜市）\n"
        "· 成语（一飞冲天）\n· 时间词（子时、春分）",
        128, star=True); row += 1
    question_row(ws, row, "6. ★ 特别注意事项",
        "专写容易搞混的边界规则，写成可判定的硬规则。",
        "· 老X/小X/阿X、X嫂/X爷/X娘 一律人名，一个不漏\n"
        "· 排行命名（戈老大、戈老二、钱二娘）全部提取\n"
        "· 单字动物名作角色名时提取（青、燕、隼）\n"
        "·「青」是角色名要提，「青色」不提",
        110, star=True); row += 1

    # ── 三、术语分类 ──────────────────────────────────────────────
    section_header(ws, row, "三、术语分类"); row += 1
    question_row(ws, row, "7. 术语分类列表",
        "每行一类，够用即可，别过细——分类太多太细模型会摇摆。\n"
        "可直接复制示例改。",
        "武学招式、武学心法、墨家机关、NPC名、BOSS名、\n"
        "门派势力、地名建筑、道具物品、圣物法宝、\n"
        "代币货币、资源材料、战斗BUFF、战斗机制、\n"
        "UI系统、玩法机制、任务名",
        160); row += 1

    # ── 四、翻译策略 ──────────────────────────────────────────────
    section_header(ws, row, "四、翻译策略（选填）"); row += 1
    question_row(ws, row, "8. 目标语言",
        "默认中→英。",
        "英文", 30); row += 1
    question_row(ws, row, "9. 翻译策略说明",
        "各类译法方向。无特别要求写「统一意译」即可。",
        "· 人名 → 音译（拼音）\n· 招式/心法 → 意译，传达含义\n"
        "· 地点 → 意译为主\n· 道具 → 简洁意译\n"
        "· 称谓 → 公子=Master，长老=Elder",
        110); row += 1

    # ── 五、举例 ──────────────────────────────────────────────────
    section_header(ws, row, "五、举例（最影响准确率 · 正例提召回，负例提精度）"); row += 1
    question_row(ws, row, "10. ★ 应该提取的例子",
        "贴几条原文 + 标出哪些是术语、属哪类。\n覆盖各种类型，越多越准。",
        "「万大海常年往来于沦波坊和神仙渡」\n"
        "→ 万大海(NPC名)、沦波坊(地名建筑)、神仙渡(地名建筑)\n\n"
        "「青长老说墨门弟子需恪守门规」\n"
        "→ 青长老(NPC名)、墨门(门派势力)、门规(玩法机制)",
        130, star=True); row += 1
    question_row(ws, row, "11. ★ 不该提取的例子",
        "⚠ 必填，至少 3–5 条「整句无术语」或易误判的句子。\n"
        "只给正例不给负例 → 噪音飙升。",
        "「用木炭生火，草鞋踩在石板上」→ 无术语\n"
        "「长老说弟子们不得擅自出门，游侠也不例外」\n　→ 无术语（长老/弟子/游侠是泛称）\n"
        "「他一飞冲天，大鹏展翅般冲出」→ 无术语（成语）",
        120, star=True); row += 1

    # ── 六、其他 ──────────────────────────────────────────────────
    section_header(ws, row, "六、其他补充（选填）"); row += 1
    question_row(ws, row, "12. 其他说明",
        "任何额外需要告知的信息。",
        "（如：某些章节用方言；客户要求保留繁体专名……）",
        72)

    wb.save(OUT)
    print(f"saved: {OUT}")


if __name__ == "__main__":
    main()
